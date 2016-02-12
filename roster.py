import openpyxl
import datetime


class Member:
    def __init__(self, drn, lname, fname, city, state, sex, dob):
        self.drn = drn
        self.lname = lname
        self.fname = fname
        self.city = city
        self.state = state
        self.sex = sex
        if not isinstance(dob, (datetime.date, datetime.datetime)):
            err_msg = 'Member.__init__ age requires datetime.date/datetime'
            raise ValueError(err_msg)
        if isinstance(dob, datetime.datetime):
            dob = dob.date()
        self.dob = dob

    def __str__(self):
        s = '<Member %s: %s, %s. DOB: %s, Sex: %s>'
        return s % (self.drn, self.lname, self.fname, self.dob, self.sex)

    def age(self, as_of=None):
        if as_of is None:
            as_of = datetime.date.today()
        if not isinstance(as_of, (datetime.date, datetime.datetime)):
            err_msg = 'Member.age requires datetime.date or datetime.datetime'
            raise ValueError(err_msg)
        age = as_of.year - self.dob.year
        if (as_of.month, as_of.day) < (self.dob.month, self.dob.day):
            age -= 1
        return age

    def in_age_group(self, group_min, group_max, as_of=None):
        age = self.age(as_of=as_of)
        return group_min <= age <= group_max


def urr_rowparser(row):
    """Process one row of the URR Roster file, and return a dict with keys
    appropriate for Member.__init__:
    returned dict keys: [drn, lname, fname, city, state, sex, dob]

    Return None if it's a row to be skipped."""
    # filters
    # skip the totals row and blank DRN
    if row[0].value in ('Total', None):
        return None
    # people without a sex listed cannot be assigned to an age group
    if str(row[14].value).strip().lower() not in ('m', 'f'):
        return None
    # people without a dob listed cannot be assigned to an age group
    if row[15].value in ('', None):
        return None
    # now get the data we need
    member_info = {'drn': row[0].value,
                   'lname': row[6].value,
                   'fname': row[7].value,
                   'city': row[10].value,
                   'state': row[11].value,
                   'sex': row[14].value,
                   'dob': row[15].value,
                   }
    return member_info


def members_from_excel(src_path, sheet_name=None, row_parser=urr_rowparser):
    wb = openpyxl.load_workbook(src_path, data_only=True, guess_types=True)
    if sheet_name is not None:
        sh = wb.get_sheet_by_name(sheet_name)
    else:
        sh = wb.get_sheet_by_name(wb.sheetnames[0])
    members = []
    skipped = 0
    for row in sh.iter_rows(row_offset=1):
        fields = urr_rowparser(row)
        if fields is not None:
            m = Member(**fields)
            # print(m, 'Age %i' % m.age())
            members.append(m)
        else:
            skipped += 1
    print('Members processed: %i' % len(members))
    print('Skipped: %i' % skipped)
    return members


if __name__ == '__main__':
    # TODO: create a command line argument for src_path. Required.
    # TODO: allow for filetypes besides xlsx
    # TODO: create a command line arguemnt for sheet_name. If not present, use
    # index 0
    members = members_from_excel(r'roster.xlsx')
